"""
EMFOX OMS v2.1 - Módulo de Exportación Excel (11 columnas A-K)
================================================================
Genera archivos .xlsx con formato visual EMFOX. Columnas:
A=PHOTO, B=CODE, C=ARTICULO, D=DESCRIPTION,
E=CAJAS, F=UND POR CAJA, G=TOTAL (QUANTITY group),
H=UNIT, I=TOTAL (CBM group),
J=UNIT, K=TOTAL (USD PRICE group).
"""

import io
import os
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PIL import Image as PILImage

from app.schemas import ExportRequest
from app.config import settings

# ============================================================
# PALETA DE COLORES EMFOX
# ============================================================
NAVY_BLUE = "1B2A4A"
DARK_BLUE = "1B3A5C"
HEADER_BG = "2C3E6B"
LIGHT_GREEN = "E8F5E8"
WHITE = "FFFFFF"
BLACK = "000000"

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Image constants for Excel
PHOTO_COL_WIDTH = 16
PHOTO_ROW_HEIGHT = 65
PHOTO_CELL_PX_W = 110
PHOTO_CELL_PX_H = 80

# Last column letter for 11-column layout
LAST_COL = "K"


def _resolve_image_path(photo_url: Optional[str]) -> Optional[str]:
    """Resolve a photo URL to an absolute filesystem path."""
    if not photo_url:
        return None
    rel = photo_url.lstrip("/")
    path = Path(rel)
    if path.exists():
        return str(path)
    path = Path(settings.upload_dir).parent / rel
    if path.exists():
        return str(path)
    return None


def _prepare_image_for_excel(image_path: str) -> Optional[XlImage]:
    """Load, resize, and return an openpyxl Image object."""
    try:
        img = PILImage.open(image_path)
        img.thumbnail((PHOTO_CELL_PX_W, PHOTO_CELL_PX_H), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.convert("RGB").save(buf, format="JPEG", quality=85)
        buf.seek(0)
        xl_img = XlImage(buf)
        xl_img.width = img.width
        xl_img.height = img.height
        return xl_img
    except Exception as e:
        print(f"[EXCEL] Error preparando imagen {image_path}: {e}")
        return None


def generate_emfox_excel(data: ExportRequest) -> io.BytesIO:
    """
    Genera archivo Excel con formato EMFOX v2.1 (11 columnas A-K).
    Incluye UND POR CAJA, alineación centrada para ORIGIN/FROM/TO/PAYMENT,
    y alineación derecha para totales.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "LISTA DE PRODUCTOS"

    # Column widths (11 columns A-K)
    col_widths = {
        "A": PHOTO_COL_WIDTH, "B": 12, "C": 14, "D": 20,
        "E": 10, "F": 15, "G": 10, "H": 10, "I": 10, "J": 12, "K": 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    row = 1

    # ============================================================
    # CORPORATE HEADER
    # ============================================================
    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    cell = ws[f"A{row}"]
    cell.value = "EMFOX YIWU TRADE CO., LTD"
    cell.font = Font(name="Times New Roman", size=22, bold=True, color=NAVY_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    ws.row_dimensions[row].height = 40
    row += 1

    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    cell = ws[f"A{row}"]
    cell.value = "1229, 12TH FLOOR, BLOCK A, CHOUYIN BUILDING, NO. 188 SHANGCHENG AVENUE, FINANCIAL AND BUSINESS DISTRICT, FUTIAN STREET, YIWU CITY"
    cell.font = Font(name="Arial", size=7, color=BLACK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1

    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    cell = ws[f"A{row}"]
    cell.value = "TELE:0086-198-49046243   CONTACTO: JOMEINI"
    cell.font = Font(name="Arial", size=9, color=BLACK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1

    # ============================================================
    # TITLE BAR
    # ============================================================
    ws.merge_cells(f"A{row}:{LAST_COL}{row}")
    cell = ws[f"A{row}"]
    cell.value = "LISTA DE PRODUCTOS"
    cell.font = Font(name="Arial", size=20, bold=True, italic=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    ws.row_dimensions[row].height = 45
    row += 1

    # ============================================================
    # CONSIGNEE DATA
    # ============================================================
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws[f"A{row}"]
    cell.value = f"CONSIGNEE: {data.consignee}\nRUC: {data.ruc}\nDIRECCION: {data.direccion}"
    cell.font = Font(name="Arial", size=9, bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells(f"I{row}:{LAST_COL}{row}")
    cell = ws[f"I{row}"]
    export_date = data.date or date.today().strftime("%d/%m/%Y")
    cell.value = f"DATE: {export_date}"
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 40
    row += 1

    # ORIGIN + PAYMENT (center aligned)
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "ORIGIN PRODUCTS : CHINA"
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"G{row}:{LAST_COL}{row}")
    cell = ws[f"G{row}"]
    cell.value = f"PAYMENT TERM: {data.payment_term}"
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # FROM + TO (center aligned)
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = f"FROM: {data.origin}"
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"G{row}:{LAST_COL}{row}")
    cell = ws[f"G{row}"]
    cell.value = f"TO: {data.destination}"
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # ============================================================
    # TABLE HEADERS (2 rows with merge) - 11 columns
    # ============================================================
    header_font = Font(name="Arial", size=9, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_row1 = row

    # Single-column headers that span 2 rows
    for col, text in {"A": "PHOTO", "B": "CODE", "C": "ARTICULO", "D": "DESCRIPTION"}.items():
        cell = ws[f"{col}{header_row1}"]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN_BORDER
        ws.merge_cells(f"{col}{header_row1}:{col}{header_row1 + 1}")

    # Group headers: QUANTITY (E:G), CBM (H:I), USD PRICE (J:K)
    for cols, text in [("E:G", "QUANTITY"), ("H:I", "CBM"), ("J:K", "USD PRICE")]:
        c1, c2 = cols.split(":")
        ws.merge_cells(f"{c1}{header_row1}:{c2}{header_row1}")
        cell = ws[f"{c1}{header_row1}"]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN_BORDER

    ws.row_dimensions[header_row1].height = 22
    row += 1

    # Sub-headers row
    sub_headers = {
        "E": "CAJAS", "F": "UND POR CAJA", "G": "TOTAL",
        "H": "UNIT", "I": "TOTAL",
        "J": "UNIT", "K": "TOTAL",
    }
    for col, text in sub_headers.items():
        cell = ws[f"{col}{row}"]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 22
    row += 1

    # ============================================================
    # PRODUCT DATA ROWS (WITH IMAGES)
    # ============================================================
    data_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    total_general_usd = 0.0
    total_general_cbm = 0.0
    total_general_cajas = 0
    total_general_qty = 0

    for i, product in enumerate(data.products):
        row_fill = PatternFill(
            start_color=LIGHT_GREEN if i % 2 == 0 else WHITE,
            end_color=LIGHT_GREEN if i % 2 == 0 else WHITE,
            fill_type="solid",
        )

        ws.row_dimensions[row].height = PHOTO_ROW_HEIGHT

        # A: PHOTO
        cell = ws[f"A{row}"]
        cell.fill = row_fill
        cell.border = THIN_BORDER
        cell.alignment = center_align

        photo_url = product.crop_url or product.photo_url
        img_path = _resolve_image_path(photo_url)
        if img_path:
            xl_img = _prepare_image_for_excel(img_path)
            if xl_img:
                ws.add_image(xl_img, f"A{row}")
            else:
                cell.value = "[foto]"
        else:
            cell.value = "[foto]" if photo_url else ""

        # B: CODE
        cell = ws[f"B{row}"]
        cell.value = product.code
        cell.font = data_font
        cell.alignment = center_align
        cell.fill = row_fill
        cell.border = THIN_BORDER

        # C: ARTICULO
        cell = ws[f"C{row}"]
        cell.value = product.articulo
        cell.font = data_font
        cell.alignment = center_align
        cell.fill = row_fill
        cell.border = THIN_BORDER

        # D: DESCRIPTION
        cell = ws[f"D{row}"]
        cell.value = product.description
        cell.font = data_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = row_fill
        cell.border = THIN_BORDER

        # E: CAJAS
        cell = ws[f"E{row}"]
        cell.value = product.quantity_cajas or ""
        cell.font = data_font
        cell.alignment = center_align
        cell.fill = row_fill
        cell.border = THIN_BORDER

        # F: UND POR CAJA
        cell = ws[f"F{row}"]
        cell.value = product.quantity_und_por_caja or ""
        cell.font = data_font
        cell.alignment = center_align
        cell.fill = row_fill
        cell.border = THIN_BORDER

        # G: TOTAL QTY
        cell = ws[f"G{row}"]
        cell.value = product.quantity_total
        cell.font = data_font
        cell.alignment = center_align
        cell.fill = row_fill
        cell.border = THIN_BORDER

        # H: CBM UNIT
        cell = ws[f"H{row}"]
        cell.value = product.cbm_unit
        cell.font = data_font
        cell.alignment = center_align
        cell.fill = row_fill
        cell.border = THIN_BORDER

        # I: CBM TOTAL
        cell = ws[f"I{row}"]
        cell.value = product.cbm_total
        cell.font = data_font
        cell.alignment = center_align
        cell.fill = row_fill
        cell.border = THIN_BORDER

        # J: USD UNIT
        cell = ws[f"J{row}"]
        cell.value = f"${product.precio_unitario_usd:.2f}"
        cell.font = data_font
        cell.alignment = center_align
        cell.fill = row_fill
        cell.border = THIN_BORDER

        # K: USD TOTAL
        cell = ws[f"K{row}"]
        cell.value = f"${product.total_usd:,.2f}"
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = right_align
        cell.fill = row_fill
        cell.border = THIN_BORDER

        total_general_usd += product.total_usd
        total_general_cbm += product.cbm_total
        total_general_cajas += (product.quantity_cajas or 0)
        total_general_qty += (product.quantity_total or 0)
        row += 1

    # ============================================================
    # TOTALS ROW (right-aligned values)
    # ============================================================
    row += 1
    totals_font = Font(name="Arial", size=11, bold=True, color=WHITE)
    totals_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    totals_center = Alignment(horizontal="center", vertical="center")
    totals_right = Alignment(horizontal="right", vertical="center")

    # A-D merged: "TOTALES"
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "TOTALES"
    cell.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    cell.fill = totals_fill
    cell.alignment = totals_center
    cell.border = THIN_BORDER

    # E: Total cajas
    cell = ws[f"E{row}"]
    cell.value = total_general_cajas
    cell.font = totals_font
    cell.fill = totals_fill
    cell.alignment = totals_right
    cell.border = THIN_BORDER

    # F: empty
    cell = ws[f"F{row}"]
    cell.fill = totals_fill
    cell.border = THIN_BORDER

    # G: Total qty
    cell = ws[f"G{row}"]
    cell.value = total_general_qty
    cell.font = totals_font
    cell.fill = totals_fill
    cell.alignment = totals_right
    cell.border = THIN_BORDER

    # H: empty
    cell = ws[f"H{row}"]
    cell.fill = totals_fill
    cell.border = THIN_BORDER

    # I: Total CBM
    cell = ws[f"I{row}"]
    cell.value = f"{total_general_cbm:.2f} m\u00b3"
    cell.font = totals_font
    cell.fill = totals_fill
    cell.alignment = totals_right
    cell.border = THIN_BORDER

    # J: empty
    cell = ws[f"J{row}"]
    cell.fill = totals_fill
    cell.border = THIN_BORDER

    # K: Total USD
    cell = ws[f"K{row}"]
    cell.value = f"${total_general_usd:,.2f}"
    cell.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    cell.fill = totals_fill
    cell.alignment = totals_right
    cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 30

    # ============================================================
    # PAGE SETUP — fit everything on one page for PDF export
    # ============================================================
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Tight margins (inches) — maximize printable area
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.3, bottom=0.3,
        header=0.1, footer=0.1,
    )

    # Print area covers all content
    ws.print_area = f"A1:{LAST_COL}{row}"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
